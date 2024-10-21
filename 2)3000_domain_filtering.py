import time
import requests
import openpyxl
import concurrent.futures

"""
Fleet filter

This is a real project for one of my customers that had a big list of 
different flight companies' sites and wanted to find the ones that were
specialized for charters and fleets. So, this script goes to every site
provided in the CSV file and filtering them to the ones that specialized 
on fleets and charters and provides the list containing them.

Auther: Erik Badalyan

"""


def get_urls_(url_):
    wrkbk = openpyxl.load_workbook(url_)
    sh = wrkbk.active
    csv_urls_list = []
    for row in range(1, sh.max_row + 1):
        csv_urls_list.append(sh.cell(row=row, column=1).value)
    return csv_urls_list


def new_data_(url__: str) -> None:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36'
                      ' (KHTML, like Gecko) Chrome/108.0.0.0 Mobile Safari/537.36',
    }
    try:
        url__ = f'http://www.{url__}'
        global idx
        idx += 1
        print(f'{idx}) {url__}')
        req = requests.get(url__, headers=headers)
        print(str(req))
        if (('200' in str(req)) and ('fleet' in req.text or 'Fleet'
                                     in req.text or 'Charter' in req.text
                                     or 'charter' in req.text)) \
                and ('part' not in req.text or 'Part' not in req.text):
            new_data.append(url__)
    except Exception:
        print("Not working")
    return None


if __name__ == "__main__":
    start = time.time()
    new_data = []
    idx = 0
    with concurrent.futures.ThreadPoolExecutor() as executor:
        url = r"C:\Users\erikb\PycharmProjects\Python_Aua\New folder\Web Scraping\Roman_Test_Little.xlsx"
        urls = get_urls_(url)
        executor.map(new_data_, urls)
    print('new data - ', list(filter(lambda true_url: true_url, new_data)), end='\n\n')
    print('new data size is: ', len(list(filter(lambda true_url: true_url, new_data))), end='\n\n')
    end = time.time()
    print(f'It takes {round(end-start, 2)} seconds')
